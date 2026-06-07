import io
from django.shortcuts import get_object_or_404
from django.contrib.admin.views.decorators import staff_member_required
from django.http import HttpResponse
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, Alignment
from apps.students.models import Student
from apps.grades.models import Grade
from apps.assignments.models import Submission
from apps.academics.models import Course

@staff_member_required
def student_report_pdf(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=72)
    styles = getSampleStyleSheet()
    title_style = styles['Title']
    heading_style = styles['Heading2']
    
    story = []
    
    story.append(Paragraph(f"Student Performance Report: {student.user.get_full_name()}", title_style))
    story.append(Spacer(1, 0.2*inch))
    
    info_data = [
        ["Roll Number", student.roll_number],
        ["Enrollment", student.enrollment_number],
        ["Department", student.department.name if student.department else "-"],
        ["Batch", student.batch.year if student.batch else "-"],
        ["Semester", student.semester],
        ["Current CGPA", student.current_cgpa],
    ]
    info_table = Table(info_data, colWidths=[2*inch, 3*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (0,-1), colors.lightgrey),
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 0.3*inch))
    
    grades = Grade.objects.filter(student=student).select_related('course').order_by('semester', 'course__code')
    if grades:
        story.append(Paragraph("Academic Grades", heading_style))
        grade_data = [["Course", "Exam Type", "Marks", "Max Marks", "Percentage"]]
        for g in grades:
            grade_data.append([
                g.course.code,
                g.get_exam_type_display(),
                g.marks_obtained,
                g.maximum_marks,
                f"{g.percentage:.1f}%"
            ])
        grade_table = Table(grade_data, colWidths=[1.5*inch, 1.5*inch, 0.8*inch, 0.8*inch, 0.8*inch])
        grade_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.grey),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('GRID', (0,0), (-1,-1), 1, colors.black),
        ]))
        story.append(grade_table)
        story.append(Spacer(1, 0.3*inch))
    
    submissions = Submission.objects.filter(student=student).select_related('assignment')
    if submissions:
        story.append(Paragraph("Assignment Submissions", heading_style))
        assign_data = [["Assignment", "Course", "Submitted On", "Marks", "Late"]]
        for sub in submissions:
            assign_data.append([
                sub.assignment.title,
                sub.assignment.course.code,
                sub.submitted_at.strftime("%Y-%m-%d"),
                f"{sub.marks_obtained}/{sub.assignment.total_marks}" if sub.marks_obtained is not None else "Not graded",
                "Yes" if sub.is_late else "No"
            ])
        assign_table = Table(assign_data, colWidths=[2*inch, 1*inch, 1.2*inch, 0.8*inch, 0.5*inch])
        assign_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.grey),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('GRID', (0,0), (-1,-1), 1, colors.black),
        ]))
        story.append(assign_table)
    
    doc.build(story)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="student_{student.roll_number}_report.pdf"'
    return response

@staff_member_required
def class_performance_excel(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    enrollments = course.enrollment_set.filter(status='active').select_related('student__user')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{course.code}_performance"
    
    headers = ['Roll Number', 'Student Name', 'Semester', 'CGPA', 'Internal Marks', 'External Marks', 'Total']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    row_num = 2
    for enrollment in enrollments:
        student = enrollment.student
        grades = Grade.objects.filter(student=student, course=course)
        internal_total = sum(g.marks_obtained for g in grades if g.exam_type.startswith('internal'))
        external = sum(g.marks_obtained for g in grades if g.exam_type == 'external')
        total = internal_total + external
        
        ws.cell(row=row_num, column=1, value=student.roll_number)
        ws.cell(row=row_num, column=2, value=student.user.get_full_name())
        ws.cell(row=row_num, column=3, value=student.semester)
        ws.cell(row=row_num, column=4, value=student.current_cgpa)
        ws.cell(row=row_num, column=5, value=internal_total)
        ws.cell(row=row_num, column=6, value=external)
        ws.cell(row=row_num, column=7, value=total)
        row_num += 1
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{course.code}_performance.xlsx"'
    wb.save(response)
    return response